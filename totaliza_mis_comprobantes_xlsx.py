#!/usr/bin/python3

import openpyxl
import os
import glob
import sys
import itertools


def normpath(path=''):
    return os.path.expanduser(os.path.normpath(path))


def valida_parametros():
    if len(sys.argv) > 2:
        print('Cantidad de parametros incorrecta')
        sys.exit()
    wb_path = ''
    if len(sys.argv) == 2:
        wb_path = sys.argv[1]
        return normpath(wb_path.rstrip())
    return False


def seParecen(numero1, numero2, diferenciaDePesosTolerable):
    if (abs(float(numero1) - float(numero2)) > diferenciaDePesosTolerable):
        return False
    return True


def none_floater(variable):
    if variable is None:
        return 0
    return variable


def none_crusher(variable):
    if variable is None:
        return ''
    return variable


def alicuotas_verificadas(neto=float(0.0), iva=float(0.0), alicuotas=[float(21), float(27), float(10.5)]):
    combinaciones_de_alicuotas = list()
    for x in range(1, len(alicuotas)):
        for combinacion in list(itertools.combinations(alicuotas, x)):
            combinaciones_de_alicuotas.append(combinacion)
    for combinacion in combinaciones_de_alicuotas:
        acumulador = float(0.0)
        for alicuota in combinacion:
            acumulador += (neto * alicuota / 100)
        if seParecen(acumulador, iva, 0.10):
            return combinacion
    return False


def loguear(mensaje, url_log='log.txt'):
    with open(url_log, 'a') as ar_logs:
        ar_logs.write(mensaje)
        return None


def obtiene_n_fila_ultimo_dato(wb_url, n_fila_dato_inicial='1', col_testigo='A'):
    wb = openpyxl.load_workbook(wb_url)
    ws = wb.active
    n_fila_ultimo_dato = int(n_fila_dato_inicial) - 1
    while ws[str(col_testigo) + str(n_fila_ultimo_dato + 1)].value is not None:
        n_fila_ultimo_dato += 1
    wb.close()
    return n_fila_ultimo_dato


def verifica_alicuotas_xlsx(wb_url,
                            columna_neto='',
                            columna_iva='',
                            fila_dato_inicial='1',
                            col_testigo='A'):
    sin_errores = True
    fila_ultimo_dato = obtiene_n_fila_ultimo_dato(wb_url, n_fila_dato_inicial=fila_dato_inicial,
                                                  col_testigo=col_testigo)
    if fila_ultimo_dato < int(fila_dato_inicial):
        loguear('El archivo ' + str(os.path.split(wb_url)[1]) + ' se encuentra vacío')
        return False
    wb = openpyxl.load_workbook(wb_url)
    ws = wb.active
    for fila in range(int(fila_dato_inicial), fila_ultimo_dato + 1):
        if not alicuotas_verificadas(
                neto=float(none_floater(ws[str(columna_neto) + str(fila)].value)),
                iva=float(none_floater(ws[str(columna_iva) + str(fila)].value))):
            sin_errores = False
            loguear('NO se verifica combinatoria de alicuotas posible para [' + str(columna_neto) + str(
                fila) + '] en ' + str(os.path.split(wb_url)[1])+'\n')
            print('NO se verifica combinatoria de alicuotas posible para [' + str(columna_neto) + str(
                fila) + '] en ' + str(os.path.split(wb_url)[1]))
    wb.close()
    return sin_errores


def totaliza_xlsx(wb_url,
                  columnas_que_importan=[],
                  fila_dato_inicial='1',
                  sufijo='',
                  prefijo='',
                  col_testigo='A',
                  col_denominacion='I',
                  col_tipo_comprobante='B'):
    fila_ultimo_dato = obtiene_n_fila_ultimo_dato(wb_url, n_fila_dato_inicial=fila_dato_inicial,
                                                  col_testigo=col_testigo)
    wb = openpyxl.load_workbook(wb_url)
    ws = wb.active
    if fila_ultimo_dato < int(fila_dato_inicial):
        wb.save(wb_url.rstrip('.xlsx') + sufijo + '.xlsx')
        return False
    ws[str(col_denominacion) + str(fila_ultimo_dato + 2)] = 'TOTALES :'
    for col in columnas_que_importan:
        total = float(0.0)
        for fila in range(int(fila_dato_inicial), fila_ultimo_dato + 1):
            signo = 1
            if ws[str(col) + str(fila)].value is not None:
                if any(tipo_comprobante.lower() in str(ws[str(col_tipo_comprobante) + str(fila)].value).lower()
                       for tipo_comprobante in ['crédito', 'credito', 'cred']):
                    signo = -1
                total += (float(ws[str(col) + str(fila)].value) * signo)
        ws[str(col) + str(fila_ultimo_dato + 2)] = round(total, 2)
    head, tail = os.path.split(wb_url)
    wb_url = os.path.join(head, prefijo + tail)
    wb.save(wb_url.rstrip('.xlsx') + sufijo + '.xlsx')
    wb.close()
    return True


wb_path = valida_parametros()
if not wb_path:
    wb_path = normpath(input('Arrastre el dir:').rstrip())
    if not os.path.isdir(wb_path):
        print('El directorio no existe.')
        sys.exit()

for archivo in sorted(glob.glob(os.path.join(wb_path, '*.xlsx'))):
    print('verificando alicuotas para ' + str(os.path.split(archivo)[1]).ljust(50) + ' :', end='')
    if verifica_alicuotas_xlsx(archivo, columna_neto='L', columna_iva='O', fila_dato_inicial='3'):
        print('OK', end='')
    print('')

for archivo in sorted(glob.glob(os.path.join(wb_path, '*.xlsx'))):
    output = 'totalizado ' + str(os.path.split(archivo)[1]).ljust(50)
    if not totaliza_xlsx(archivo, columnas_que_importan=['L', 'M', 'N', 'O', 'P'], fila_dato_inicial='3',
                         prefijo='totalizado_'):
        output = 'No se pudo totalizar ' + str(os.path.split(archivo)[1]) + '. Posiblemente no contiene datos.'
    print(output)
