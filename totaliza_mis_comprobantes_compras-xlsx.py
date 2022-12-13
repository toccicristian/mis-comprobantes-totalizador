#!/usr/bin/python3

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import alignment
import os
import glob
import sys
import itertools
import configparser
import clases.orden_columnas

configuracion=configparser.ConfigParser()
configuracion.read('config.ini')
ordcol=clases.orden_columnas.Orden_columnas(pv=configuracion.get('Iva Compras','punto de ventas'),
                      n_comp=configuracion.get('Iva Compras','numero de comprobante'),
                      t_comp=configuracion.get('Iva Compras','tipo de comprobante'),
                      denominacion=configuracion.get('Iva Compras','denominacion'),
                      n_documento=configuracion.get('Iva Compras','numero de documento'),
                      t_documento=configuracion.get('Iva Compras','tipo de documento'),
                      t_cambio=configuracion.get('Iva Compras','tipo de cambio'),
                      neto=configuracion.get('Iva Compras','neto'),
                      neto_no_g=configuracion.get('Iva Compras','neto no gravado'),
                      exento=configuracion.get('Iva Compras','exento'),
                      iva=configuracion.get('Iva Compras','iva'),
                      total=configuracion.get('Iva Compras','total'))


def normpath(path=''):
    return os.path.expanduser(os.path.normpath(path))


def valida_parametros():
    if len(sys.argv) > 2:
        print('Cantidad de parametros incorrecta')
        sys.exit()
    wb_path = ''
    if len(sys.argv) == 2:
        wb_path = sys.argv[1]
        return normpath(wb_path.rstrip())
    return False


def seParecen(numero1, numero2, diferenciaDePesosTolerable):
    if abs(float(numero1) - float(numero2)) > diferenciaDePesosTolerable:
        return False
    return True


def none_floater(variable):
    if variable is None:
        return 0
    return variable


def none_crusher(variable):
    if variable is None:
        return str('')
    return variable


def alicuotas_verificadas(neto=float(0.0), iva=float(0.0), alicuotas=[float(21), float(27), float(10.5)]):
    combinaciones_de_alicuotas = list()
    for x in range(1, len(alicuotas)):
        for combinacion in list(itertools.combinations(alicuotas, x)):
            combinaciones_de_alicuotas.append(combinacion)
    for combinacion in combinaciones_de_alicuotas:
        acumulador = float(0.0)
        for alicuota in combinacion:
            acumulador += (neto * alicuota / 100)
        if seParecen(acumulador, iva, 0.10):
            return combinacion
    return False


def loguear(mensaje, url_log='log.txt'):
    with open(url_log, 'a') as ar_logs:
        ar_logs.write(mensaje)
        return None


def obtiene_n_fila_ultimo_dato(wb_url, n_fila_dato_inicial='1', col_testigo='A'):
    wb = openpyxl.load_workbook(wb_url)
    ws = wb.active
    n_fila_ultimo_dato = int(n_fila_dato_inicial) - 1
    while ws[str(col_testigo) + str(n_fila_ultimo_dato + 1)].value is not None:
        n_fila_ultimo_dato += 1
    wb.close()
    return n_fila_ultimo_dato


def verifica_alicuotas_xlsx(wb_url,
                            orden=clases.orden_columnas.Orden_columnas,
                            fila_dato_inicial='1',
                            col_testigo='A'):
    sin_errores = True
    fila_ultimo_dato = obtiene_n_fila_ultimo_dato(wb_url, n_fila_dato_inicial=fila_dato_inicial,
                                                  col_testigo=col_testigo)
    if fila_ultimo_dato < int(fila_dato_inicial):
        loguear('El archivo ' + str(os.path.split(wb_url)[1]) + ' se encuentra vacío')
        return False
    wb = openpyxl.load_workbook(wb_url)
    ws = wb.active
    for fila in range(int(fila_dato_inicial), fila_ultimo_dato + 1):
        if not alicuotas_verificadas(
                neto=float(none_floater(ws[str(orden.neto) + str(fila)].value)),
                iva=float(none_floater(ws[str(orden.iva) + str(fila)].value))):
            sin_errores = False
            loguear('NO se verifica combinatoria de alicuotas posible para [' + str(orden.neto) + str(
                fila) + '] en ' + str(os.path.split(wb_url)[1]) + '\n')
            print('NO se verifica combinatoria de alicuotas posible para [' + str(orden.neto) + str(
                fila) + '] en ' + str(os.path.split(wb_url)[1]))
    wb.close()
    return sin_errores


def celda_fnorm(celda):
    if celda.value is None:
        return float(0.0)
    return float(str(celda.value).replace(',', '.'))


def corrige_valores_compra(wb_url, fila_dato_inicial='3', col_testigo='A', orden=clases.orden_columnas.Orden_columnas):
    wb = openpyxl.load_workbook(wb_url)
    ws = wb.active
    if obtiene_n_fila_ultimo_dato(wb_url, n_fila_dato_inicial=fila_dato_inicial, col_testigo=col_testigo) < int(fila_dato_inicial):
        wb.save(wb_url.rstrip('.xlsx') + '_corregido-enblanco.xlsx')
        return False

    col_tipo_comp = orden.t_comp
    col_tcambio = orden.t_cambio
    col_netog = orden.neto
    col_nog = orden.neto_no_g
    col_op_ex = orden.exento
    col_iva = orden.iva
    col_total = orden.total
    ws[str(orden.n_comp)+'2'] = 'N. Comp.'

    for fila in range(int(fila_dato_inicial), obtiene_n_fila_ultimo_dato(wb_url, n_fila_dato_inicial=fila_dato_inicial, col_testigo=col_testigo) + 1):
        alic=21
        if alicuotas_verificadas(neto=celda_fnorm(ws[col_netog + str(fila)]),iva=celda_fnorm(ws[col_iva + str(fila)])):
            alic=alicuotas_verificadas(neto=celda_fnorm(ws[col_netog + str(fila)]),iva=celda_fnorm(ws[col_iva + str(fila)]))[0]

        ws[col_iva + str(fila)] = round(celda_fnorm(ws[col_netog + str(fila)]) * alic / 100,2)
        ws[col_nog + str(fila)] = round(celda_fnorm(ws[col_total + str(fila)]) -
                                        celda_fnorm(ws[col_netog + str(fila)]) -
                                        celda_fnorm(ws[col_op_ex + str(fila)]) -
                                        celda_fnorm(ws[col_iva + str(fila)]), 2)

        if float(ws[col_nog + str(fila)].value) < 0.0:
            ws[col_nog + str(fila)] = round(0.0, 2)

        for col in [col_netog,col_nog,col_op_ex,col_iva,col_total]: # convierto a tipo de cambio:
            ws[col + str(fila)] = round(celda_fnorm(ws[col + str(fila)])*celda_fnorm(ws[col_tcambio + str(fila)]),2)

        if 'factura c' in str(ws[col_tipo_comp + str(fila)].value).lower():
            ws[col_netog + str(fila)] = float(0.0)
            ws[col_nog + str(fila)] = float(0.0)
            ws[col_op_ex + str(fila)] = float(0.0)
            ws[col_iva + str(fila)] = float(0.0)

    wb.save(wb_url.rstrip('.xlsx') + '_corregido.xlsx')
    wb.close()
    return wb_url.rstrip('.xlsx') + '_corregido.xlsx'


def totaliza_xlsx(wb_url,
                  orden=clases.orden_columnas.Orden_columnas,
                  fila_dato_inicial='1',
                  sufijo='',
                  prefijo='',
                  col_testigo='A'):
    fila_ultimo_dato = obtiene_n_fila_ultimo_dato(wb_url, n_fila_dato_inicial=fila_dato_inicial,
                                                  col_testigo=col_testigo)

    columnas_que_importan=[orden.neto, orden.neto_no_g, orden.exento, orden.iva, orden.total]
    wb = openpyxl.load_workbook(wb_url)
    ws = wb.active
    if fila_ultimo_dato < int(fila_dato_inicial):
        wb.save(wb_url.rstrip('.xlsx') + sufijo + '.xlsx')
        return False

    ws[str(orden.denominacion) + str(fila_ultimo_dato + 2)] = 'TOTALES :'
    for col in columnas_que_importan:
        total = float(0.0)
        for fila in range(int(fila_dato_inicial), fila_ultimo_dato + 1):
            signo = 1
            if ws[str(col) + str(fila)].value is not None:
                if any(tipo_comprobante.lower() in str(ws[str(orden.t_comp) + str(fila)].value).lower()
                       for tipo_comprobante in ['crédito', 'credito', 'cred']):
                    signo = -1
                total += (float(ws[str(col) + str(fila)].value) * signo)
        ws[str(col) + str(fila_ultimo_dato + 2)] = round(total, 2)
    head, tail = os.path.split(wb_url)
    wb_url = os.path.join(head, prefijo + tail)
    wb.save(wb_url.rstrip('.xlsx') + sufijo + '.xlsx')
    wb.close()
    return True


def formatea_matriz_numeros(ws, ncol_i=int(), ncol_f=int(), nfila_i=int(), nfila_f=int(), formato='#,##0.00'):
    for ncol in range(ncol_i, ncol_f):
        for nfila in range(nfila_i, nfila_f):
            ws[get_column_letter(ncol) + str(nfila)].number_format = formato
    return True


def ajusta_columnas(hoja, cushion=int(2), fila_inicial=int(2)):
    for ncol in range(hoja.min_column, hoja.max_column + 1):
        lista_celdas = list()
        for nfila in range(fila_inicial, hoja.max_row + 1):
            lista_celdas.append(hoja[get_column_letter(ncol) + str(nfila)].value)
        max_w = 0
        for item in lista_celdas:
            if len(str(none_crusher(item))) > max_w:
                max_w = len(str(none_crusher(item)))
        hoja.column_dimensions[get_column_letter(ncol)].width = max_w + cushion


def ajusta_columna(hoja, ncol, cushion=int(8), fila_inicial=int(2)):
    lista_celdas = list()
    for nfila in range(fila_inicial, hoja.max_row + 1):
        lista_celdas.append(hoja[get_column_letter(ncol) + str(nfila)].value)
    max_w = 0
    for item in lista_celdas:
        if len(str(none_crusher(item))) > max_w:
            max_w = len(str(none_crusher(item)))
    hoja.column_dimensions[get_column_letter(ncol)].width = max_w + cushion


def corrige_nombres_campo(hoja, fila_titulos=int(2),orden=clases.orden_columnas.Orden_columnas):
    col_pv = f'{orden.pv}'
    col_n_comp = f'{orden.n_comp}'
    col_tipo_doc = f'{orden.t_documento}'
    col_n_doc = f'{orden.n_documento}'
    col_t_cambio = f'{orden.t_cambio}'
    col_neto_g = f'{orden.neto}'
    col_neto_no_g = f'{orden.neto_no_g}'
    col_exento = f'{orden.exento}'
    hoja[col_pv + str(fila_titulos)] = 'P.Venta'
    hoja[col_n_comp + str(fila_titulos)] = 'N.Comp.'
    hoja[col_tipo_doc + str(fila_titulos)] = 'Tipo Doc.'
    hoja[col_n_doc + str(fila_titulos)] = 'N. Doc.'
    hoja[col_t_cambio + str(fila_titulos)] = 'T.Cambio'
    hoja[col_neto_g + str(fila_titulos)] = 'Neto G.'
    hoja[col_neto_no_g + str(fila_titulos)] = 'Neto no G.'
    hoja[col_exento + str(fila_titulos)] = 'Exento'


wb_path = valida_parametros()
if not wb_path:
    wb_path = normpath(input('Arrastre el dir:').rstrip())
    if not os.path.isdir(wb_path):
        print('El directorio no existe.')
        sys.exit()

razon_social = str(input('RAZON SOCIAL :'))
cuit = str(input('CUIT         :'))
titulo = str(input('TITULO       :'))
for archivo in sorted(glob.glob(os.path.join(wb_path, '*.xlsx'))):
    print('verificando alicuotas para ' + str(os.path.split(archivo)[1]).ljust(50) + ' :', end='')
    if verifica_alicuotas_xlsx(archivo, ordcol, fila_dato_inicial='3'):
        print('OK', end='')
    print('')

for archivo in sorted(glob.glob(os.path.join(wb_path, '*.xlsx'))):
    output = 'totalizado ' + str(os.path.split(archivo)[1]).ljust(50)
    url_corregido = corrige_valores_compra(wb_url=archivo, fila_dato_inicial='3', col_testigo='A',orden=ordcol)
    if not totaliza_xlsx(url_corregido, orden=ordcol, fila_dato_inicial='3',
                         prefijo=''):
        output = 'No se pudo totalizar ' + str(os.path.split(archivo)[1]) + '. Posiblemente no contiene datos.'
    print(output)
    wb = openpyxl.load_workbook(url_corregido)
    ws = wb.active
    corrige_nombres_campo(ws,orden=ordcol)
    for celdas_mergeadas in list(ws.merged_cells):
        ws.unmerge_cells(range_string=str(celdas_mergeadas))
    ws['A1'] = razon_social
    ws.cell(1, 1).alignment = openpyxl.styles.Alignment(horizontal='left')
    ws['C1'] = 'CUIT :'+cuit
    ws.cell(1, 3).alignment = openpyxl.styles.Alignment(horizontal='left')
    ws['I1'] = titulo
    ws.cell(1, 9).alignment = openpyxl.styles.Alignment(horizontal='left')
    ws.delete_cols(5, 2)
    ajusta_columnas(ws, cushion=2, fila_inicial=2)
    ajusta_columna(ws, ncol=1, cushion=1, fila_inicial=2)
    ajusta_columna(ws, ncol=7, cushion=6, fila_inicial=2)
    formatea_matriz_numeros(ws, ws.max_column - 4, ws.max_column + 1, ws.min_row + 2, ws.max_row + 3)
    for n_col in range(10, 15):
        ajusta_columna(ws, ncol=n_col, cushion=4)
    ajusta_columna(ws, ncol=ws.max_column - 2, cushion=6)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperHeight = '297mm'
    ws.page_setup.paperWidth = '210mm'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = False
    wb.save(url_corregido)
    wb.close()
