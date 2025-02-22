#!/usr/bin/python3

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import alignment
import os
import glob
import sys
import itertools
import configparser
import clases.orden_columnas


configuracion=configparser.ConfigParser()
configuracion.read('config.ini')
ordcol=clases.orden_columnas.Orden_columnas(pv=configuracion.get('Iva Ventas','punto de ventas'),
                      n_comp=configuracion.get('Iva Ventas','numero de comprobante'),
                      t_comp=configuracion.get('Iva Ventas','tipo de comprobante'),
                      denominacion=configuracion.get('Iva Ventas','denominacion'),
                      n_documento=configuracion.get('Iva Ventas','numero de documento'),
                      t_documento=configuracion.get('Iva Ventas','tipo de documento'),
                      t_cambio=configuracion.get('Iva Ventas','tipo de cambio'),
                      neto=configuracion.get('Iva Ventas','neto'),
                      neto_no_g=configuracion.get('Iva Ventas','neto no gravado'),
                      exento=configuracion.get('Iva Ventas','exento'),
                      iva=configuracion.get('Iva Ventas','iva'),
                      total=configuracion.get('Iva Ventas','total'))


def normpath(path=''):
    return os.path.expanduser(os.path.normpath(path))


def valida_parametros():
    if len(sys.argv) > 2:
        print('Cantidad de parametros incorrecta')
        sys.exit()
    wb_path = ''
    if len(sys.argv) == 2:
        wb_path = sys.argv[1]
        return normpath(wb_path.rstrip())
    return False


def seParecen(numero1, numero2, diferenciaDePesosTolerable):
    if (abs(float(numero1) - float(numero2)) > diferenciaDePesosTolerable):
        return False
    return True


def none_floater(variable):
    if variable is None:
        return 0
    return variable


def none_crusher(variable):
    if variable is None:
        return ''
    return variable


def alicuotas_verificadas(neto=float(0.0), iva=float(0.0), alicuotas=[float(21), float(27), float(10.5)]):
    combinaciones_de_alicuotas = list()
    for x in range(1, len(alicuotas)):
        for combinacion in list(itertools.combinations(alicuotas, x)):
            combinaciones_de_alicuotas.append(combinacion)
    for combinacion in combinaciones_de_alicuotas:
        acumulador = float(0.0)
        for alicuota in combinacion:
            acumulador += (neto * alicuota / 100)
        if seParecen(acumulador, iva, 0.10):
            return combinacion
    return False


def loguear(mensaje, url_log='log.txt'):
    with open(url_log, 'a') as ar_logs:
        ar_logs.write(mensaje)
        return None


def obtiene_n_fila_ultimo_dato(wb_url, n_fila_dato_inicial='1', col_testigo='A'):
    wb = openpyxl.load_workbook(wb_url)
    ws = wb.active
    n_fila_ultimo_dato = int(n_fila_dato_inicial) - 1
    while ws[str(col_testigo) + str(n_fila_ultimo_dato + 1)].value is not None:
        n_fila_ultimo_dato += 1
    wb.close()
    return n_fila_ultimo_dato


def verifica_alicuotas_xlsx(wb_url,
                            orden=clases.orden_columnas.Orden_columnas,
                            fila_dato_inicial='1',
                            col_testigo='A'):
    sin_errores = True
    fila_ultimo_dato = obtiene_n_fila_ultimo_dato(wb_url, n_fila_dato_inicial=fila_dato_inicial,
                                                  col_testigo=col_testigo)
    if fila_ultimo_dato < int(fila_dato_inicial):
        loguear('El archivo ' + str(os.path.split(wb_url)[1]) + ' se encuentra vacío')
        return False
    wb = openpyxl.load_workbook(wb_url)
    ws = wb.active
    for fila in range(int(fila_dato_inicial), fila_ultimo_dato + 1):
        if not alicuotas_verificadas(
                neto=float(none_floater(ws[str(orden.neto) + str(fila)].value)),
                iva=float(none_floater(ws[str(orden.iva) + str(fila)].value))):
            sin_errores = False
            loguear('NO se verifica combinatoria de alicuotas posible para [' + str(orden.neto) + str(
                fila) + '] en ' + str(os.path.split(wb_url)[1])+'\n')
            print('NO se verifica combinatoria de alicuotas posible para [' + str(orden.neto) + str(
                fila) + '] en ' + str(os.path.split(wb_url)[1]))
    wb.close()
    return sin_errores


def totaliza_xlsx(wb_url,
                  orden=clases.orden_columnas.Orden_columnas,
                  fila_dato_inicial='1',
                  sufijo='',
                  prefijo='',
                  col_testigo='A'):
    fila_ultimo_dato = obtiene_n_fila_ultimo_dato(wb_url, n_fila_dato_inicial=fila_dato_inicial,
                                                  col_testigo=col_testigo)

    columnas_que_importan=[orden.neto, orden.neto_no_g, orden.exento, orden.iva, orden.total]
    wb = openpyxl.load_workbook(wb_url)
    ws = wb.active
    if fila_ultimo_dato < int(fila_dato_inicial):
        wb.save(wb_url.rstrip('.xlsx') + sufijo + '.xlsx')
        return False

    ws[str(orden.denominacion) + str(fila_ultimo_dato + 2)] = 'TOTALES :'
    for col in columnas_que_importan:
        total = float(0.0)
        for fila in range(int(fila_dato_inicial), fila_ultimo_dato + 1):
            signo = 1
            if ws[str(col) + str(fila)].value is not None:
                if any(tipo_comprobante.lower() in str(ws[str(orden.t_comp) + str(fila)].value).lower()
                       for tipo_comprobante in ['crédito', 'credito', 'cred']):
                    signo = -1
                total += (float(ws[str(col) + str(fila)].value) * signo)
        ws[str(col) + str(fila_ultimo_dato + 2)] = round(total, 2)
    head, tail = os.path.split(wb_url)
    wb_url = os.path.join(head, prefijo + tail)
    wb.save(wb_url.rstrip('.xlsx') + sufijo + '.xlsx')
    wb.close()
    return True


def corrige_nombres_campo(hoja, fila_titulos=int(2)):
    hoja[ordcol.pv + str(fila_titulos)] = 'P.Venta'
    hoja[ordcol.n_comp + str(fila_titulos)] = 'N.Comp.'
    hoja[ordcol.t_comp + str(fila_titulos)] = 'Tipo Doc.'
    hoja[ordcol.n_documento + str(fila_titulos)] = 'N. Doc.'
    hoja[ordcol.t_cambio + str(fila_titulos)] = 'T.Cambio'
    hoja[ordcol.neto + str(fila_titulos)] = 'Neto G.'
    hoja[ordcol.neto_no_g + str(fila_titulos)] = 'Neto no G.'
    hoja[ordcol.exento + str(fila_titulos)] = 'Exento'


def formatea_matriz_numeros(ws, ncol_i=int(), ncol_f=int(), nfila_i=int(), nfila_f=int(), formato='#,##0.00'):
    for ncol in range(ncol_i, ncol_f):
        for nfila in range(nfila_i, nfila_f):
            ws[get_column_letter(ncol) + str(nfila)].number_format = formato
    return True


def ajusta_columnas(hoja, cushion=int(2), fila_inicial=int(2)):
    for ncol in range(hoja.min_column, hoja.max_column + 1):
        lista_celdas = list()
        for nfila in range(fila_inicial, hoja.max_row + 1):
            lista_celdas.append(hoja[get_column_letter(ncol) + str(nfila)].value)
        max_w = 0
        for item in lista_celdas:
            if len(str(none_crusher(item))) > max_w:
                max_w = len(str(none_crusher(item)))
        hoja.column_dimensions[get_column_letter(ncol)].width = max_w + cushion


def ajusta_columna(hoja, ncol, cushion=int(8), fila_inicial=int(2)):
    lista_celdas = list()
    for nfila in range(fila_inicial, hoja.max_row + 1):
        lista_celdas.append(hoja[get_column_letter(ncol) + str(nfila)].value)
    max_w = 0
    for item in lista_celdas:
        if len(str(none_crusher(item))) > max_w:
            max_w = len(str(none_crusher(item)))
    hoja.column_dimensions[get_column_letter(ncol)].width = max_w + cushion


def formatea_wb(wb_url, nombre="", cuit="", titulo=""):
    wb = openpyxl.load_workbook(wb_url)
    ws = wb.active
    corrige_nombres_campo(ws)
    for celdas_mergeadas in list(ws.merged_cells):
        ws.unmerge_cells(range_string=str(celdas_mergeadas))
    ws['A1'] = nombre
    ws.cell(1, 1).alignment = openpyxl.styles.Alignment(horizontal='left')
    ws['C1'] = 'CUIT :'+cuit
    ws.cell(1, 3).alignment = openpyxl.styles.Alignment(horizontal='left')
    ws['I1'] = titulo
    ws.cell(1, 9).alignment = openpyxl.styles.Alignment(horizontal='left')
    ws.delete_cols(5, 2)
    ajusta_columnas(ws, cushion=2, fila_inicial=2)
    ajusta_columna(ws, ncol=1, cushion=1, fila_inicial=2)
    ajusta_columna(ws, ncol=7, cushion=6, fila_inicial=2)
    formatea_matriz_numeros(ws, ws.max_column - 4, ws.max_column + 1, ws.min_row + 2, ws.max_row + 3)
    for n_col in range(10, 15):
        ajusta_columna(ws, ncol=n_col, cushion=4)
    ajusta_columna(ws, ncol=ws.max_column - 2, cushion=6)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperHeight = '297mm'
    ws.page_setup.paperWidth = '210mm'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = False
    wb.save(wb_url)
    wb.close()


wb_path = valida_parametros()
if not wb_path:
    wb_path = normpath(input('Arrastre el dir:').rstrip())
    if not os.path.isdir(wb_path):
        print('El directorio no existe.')
        sys.exit()

razon_social = str(input('RAZON SOCIAL :'))
n_cuit = str(input('CUIT         :'))
titulo_libro = str(input('TITULO       :'))


for archivo in sorted(glob.glob(os.path.join(wb_path, '*.xlsx'))):
    print('verificando alicuotas para ' + str(os.path.split(archivo)[1]).ljust(50) + ' :', end='')
    if verifica_alicuotas_xlsx(archivo, orden=ordcol, fila_dato_inicial='3'):
        print('OK', end='')
    print('')

for archivo in sorted(glob.glob(os.path.join(wb_path, '*.xlsx'))):
    output = 'totalizado ' + str(os.path.split(archivo)[1]).ljust(50)
    if not totaliza_xlsx(archivo, orden=ordcol, fila_dato_inicial='3',
                         prefijo='totalizado_'):
        output = 'No se pudo totalizar ' + str(os.path.split(archivo)[1]) + '. Posiblemente no contiene datos.'
    print(output)


for archivo in sorted(glob.glob(os.path.join(wb_path, "*.xlsx"))):
    formatea_wb(archivo, nombre=razon_social, cuit=n_cuit, titulo=titulo_libro)



